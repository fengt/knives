from openpyxl import load_workbook

# def read():

# if __name__ == "__main__":
#     read()

update_sql = "UPDATE RFD_FMS.PIXIU_UP_FREIGHT_INFO ii SET ii.MERCHANTTYPE='{a}',\
ii.PRINCIPAL='{b}',ii.NEGOTIATER='{c}',ii.DEPARTMENT='{d}' WHERE ii.DEFINITIONNAME='{name}';"

path = "/Users/ftag/Downloads/dic.xlsx"
wb = load_workbook(filename = path)
sheet = wb.get_sheet_by_name("Sheet1")
print(sheet['C2'].value)
print(sheet.max_row)
print(sheet.max_column)


txt_name = "/Users/ftag/Downloads/dic.txt"
f = open(txt_name, "a")
# for cell in list(sheet.columns)[2]:
#     if cell.value == "#N/A":
#         continue
#     else:
#         # f.write(cell.value)
#         # f.write("\n")
#         # print(cell.value)


# for i in range(sheet.nrows):
#     if i == 0:
#         continue
#     else:
#         name = sheet.cell(i, 2).value
#         list.append(name)
# print(list)



for row in sheet.rows:
    if row[2].value == "#N/A":
        continue
    else:
        # print(row[2].value, row[3].value, row[4].value, row[5].value, row[6].value)
        b = update_sql
        a = b.replace("{name}", str(row[2].value))\
        .replace('{a}', str(row[3].value))\
        .replace('{b}', str(row[4].value))\
        .replace('{c}', row[5].value)\
        .replace('{d}', row[6].value)
        # print(a)
        f.write(a)
        f.write("\n")

f.close()
    